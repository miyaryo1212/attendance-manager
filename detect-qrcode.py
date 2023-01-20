import copy
import ctypes
import datetime
import os
import time
import tkinter
import tkinter.filedialog
import tkinter.messagebox

import cv2 as cv
import numpy as np
import openpyxl
import pyautogui
from pyzbar.pyzbar import decode

# bgr color dict
color = {
    "blue": (255, 0, 0),
    "green": (0, 255, 0),
    "red": ((0, 0, 255)),
}

# optimized for my shcool
# 11-19HR, 21-29HR, 30-39HR each class has less than 45 students
# inactive=-1 / active=1
# hard-coding
groups = [
    "11",
    "12",
    "13",
    "14",
    "15",
    "16",
    "17",
    "18",
    "19",
    "21",
    "22",
    "23",
    "24",
    "25",
    "26",
    "27",
    "28",
    "29",
    "30",
    "31",
    "32",
    "33",
    "34",
    "35",
    "36",
    "37",
    "38",
    "39",
]

status = {}
for group in groups:
    status[group] = np.full(46, -1)

last_action = {}
for group in groups:
    last_action[group] = np.full(46, time.time())

active_members = {}
for group in groups:
    active_members[group] = []

last_output_time = time.time()

# gui messagebox
def msgbox(title, content):
    root = tkinter.Tk()
    root.withdraw()
    tkinter.messagebox.showinfo(title, content)

    return None


# gui filedialog
def openfile(file_type=[("", "*")]):
    root = tkinter.Tk()
    root.withdraw()
    file_path = tkinter.filedialog.askopenfilename(filetypes=file_type)

    if not file_path:
        quit()

    return file_path


# gui folderdialog
def openfolder():
    root = tkinter.Tk()
    root.withdraw()
    dir_path = tkinter.filedialog.askdirectory()

    if not dir_path:
        quit()

    return dir_path


# get capture info
def getcapinfo(cap):
    w, h, f = (
        int(cap.get(cv.CAP_PROP_FRAME_WIDTH)),
        int(cap.get(cv.CAP_PROP_FRAME_HEIGHT)),
        int(cap.get(cv.CAP_PROP_FPS)),
    )

    return w, h, f


# resize caputre
def resizecap(cap, scale=0.8):
    w, h = (
        int(cap.get(cv.CAP_PROP_FRAME_WIDTH)),
        int(cap.get(cv.CAP_PROP_FRAME_HEIGHT)),
    )
    displaysize_w, displaysize_h = pyautogui.size()
    available_displaysize_w, available_displaysize_h = (
        int(displaysize_w * scale),
        int(displaysize_h * scale),
    )

    if (w > available_displaysize_w) or (h > available_displaysize_h):
        if w >= h:
            resized_w = available_displaysize_w
            resized_h = int(h * (available_displaysize_w / w))
        else:
            resized_h = available_displaysize_h
            resized_w = int(w * (available_displaysize_h / h))
    else:
        resized_w, resized_h = w, h

    return resized_w, resized_h


# detect qrcode
def detectqrcode(frame):
    contents = []
    qrcodes = decode(frame)
    for qrcode in qrcodes:
        content = qrcode.data.decode("utf-8")
        contents.append(content)

        x, y, w, h = qrcode.rect
        cv.rectangle(frame, (x, y), (x + w, y + h), color["green"], 2)

    return contents, frame


def updatestatus(contents):
    global status, last_action

    max_active = 20

    for content in contents:
        key = str(int(content) // 100)
        if (
            time.time()
            - last_action[key][int(content) % 100]
            >= 10
        ):
            if not content in active_members[key]:
                if len(active_members[key]) <= max_active:
                    status[key][int(content) % 100] *= -1
                    last_action[key][
                        int(content) % 100
                    ] = time.time()
                    active_members[key].append(content)
                    print("Append {}".format(content))
                else:
                    print("Up to only 20 people: {}".format(content))
            else:
                last_action[key][
                    int(content) % 100
                ] = time.time()
                active_members[key].remove(content)
                print("Remove {}".format(content))

        else:
            pass

    return None


# check list and save to excel worksheet
def checklist(list, bookpath):
    global last_output_time

    last_output_time = time.time()
    workbook = openpyxl.load_workbook(bookpath)
    time_now = datetime.datetime.now().strftime("%H:%M")
    for group in list:
        worksheet = workbook[group]
        target_row = worksheet.max_row + 1
        worksheet.cell(target_row, 1).value = time_now
        counter = 0
        for i in list[group]:
            worksheet.cell(target_row, counter + 2).value = i
            counter += 1
    workbook.save(bookpath)


# read videos with cv2
def readvideo(src, bookpath):
    global status, last_action, active_members

    if src == 0:
        window_title = "marker-tracking [Camera 0]"
    else:
        window_title = "marker-tracking [{}]".format(src)
    timenow = time.strftime("%Y-%m-%d %H%M%S", time.strptime(time.ctime()))
    cap = cv.VideoCapture(src)
    if cap.isOpened() == False:
        print("Error in opening video stream of file")

    w, h, f = getcapinfo(cap)
    rw, rh = resizecap(cap)

    frame_num = 0
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break

        if frame_num % (f / 5) == 0:
            contents, frame = detectqrcode(frame)
            if not len(contents) == 0:
                updatestatus(contents)
            else:
                pass
        else:
            pass

        resized_frame = cv.resize(copy.copy(frame), dsize=(rw, rh))
        resized_frame = cv.flip(resized_frame, 1)
        cv.imshow(window_title, resized_frame)

        if time.time() - last_output_time >= 15:
            checklist(active_members, bookpath)

        if cv.waitKey(1) & 0xFF == 27:
            break

        frame_num += 1

    cap.release()
    cv.destroyAllWindows()

    return None


if __name__ == "__main__":
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(True)
    except:
        pass

    dirpath = "./output"
    if not os.path.exists(dirpath):
        os.makedirs(dirpath)
    else:
        pass

    bookname = "/{}.xlsx".format(datetime.datetime.now().strftime("%Y-%m-%d"))
    if not os.path.exists(dirpath + bookname):
        openpyxl.Workbook().save(dirpath + bookname)
        workbook = openpyxl.load_workbook(dirpath + bookname)
        for group in groups:
            workbook.create_sheet(group)
        workbook.save(dirpath + bookname)
    else:
        pass

    readvideo(0, dirpath + bookname)
